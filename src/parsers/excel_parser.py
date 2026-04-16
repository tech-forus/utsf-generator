"""
Excel / CSV Parser  (pandas-based, smart multi-pattern extraction)
===================================================================
Extracts transporter data from Excel files (.xlsx, .xls) and CSV files.

Key features:
- Multi-row header scanning: tries rows 0-9 as potential header rows
- ZONE_EXPANSION: NORTH/N -> N1,N2,N3,N4, etc.
- Zone matrix detection: score-based, accepts simplified zone names
- Pincode list detection:
    - Explicit ODA column (ODA/EDL/is_oda)
    - Delivery Y/N column (Y=served, N=ODA pincodes)
    - Pickup / COD / State columns
- Charges: key-value + two-value {v%, f_min} patterns + "X% or Rs Y" strings
- ODA charge weight-band table detection
- Does NOT bail early on first match — every sheet is tried for charges too
"""

import os
import re
import csv
from typing import Dict, List, Any, Optional, Tuple
from parsers.base_parser import BaseParser

# ── Knowledge layer (lazy-loaded to avoid circular imports) ───────────────────
_SM = None   # SmartMatcher singleton
_GV = None   # GeoValidator singleton (loaded on first parse)

# Parse audit: tracks non-exact SmartMatcher decisions for UI review
_PARSE_AUDIT: list = []

def _record_audit(match_type: str, raw: str, matched, method: str, confidence: float, source: str = ""):
    """Append one uncertain match to the parse audit trail."""
    _PARSE_AUDIT.append({
        "type":       match_type,
        "raw":        raw,
        "matched":    matched,
        "method":     method,
        "confidence": round(confidence, 2),
        "source":     source,
    })

def _get_sm():
    global _SM
    if _SM is None:
        try:
            from knowledge.smart_matcher import SmartMatcher
            _SM = SmartMatcher()
        except Exception as e:
            print(f"[ExcelParser] SmartMatcher unavailable: {e}")
            _SM = False
    return _SM if _SM is not False else None

def _get_gv():
    global _GV
    if _GV is None:
        try:
            import os as _os
            from knowledge.geo_validator import GeoValidator
            # Search for pincodes.json in several locations
            _here = _os.path.dirname(_os.path.abspath(__file__))
            _data_env = _os.environ.get("UTSF_DATA", "")
            for candidate in [
                _os.path.join(_here, "..", "..", "data", "pincodes.json"),
                _os.path.join(_here, "..", "data", "pincodes.json"),
                # UTSF_DATA may be the data/ dir (EXE) or a direct file path
                _os.path.join(_data_env, "pincodes.json") if _data_env else "",
                _data_env if _data_env and _data_env.endswith(".json") else "",
                "data/pincodes.json",
            ]:
                candidate = _os.path.normpath(candidate) if candidate else ""
                if candidate and _os.path.isfile(candidate):
                    _GV = GeoValidator(candidate)
                    break
            if _GV is None:
                print("[ExcelParser] GeoValidator: pincodes.json not found")
                _GV = False
        except Exception as e:
            print(f"[ExcelParser] GeoValidator unavailable: {e}")
            _GV = False
    return _GV if _GV is not False else None


# ---------------------------------------------------------------------------
# Zone name expansion table
# ---------------------------------------------------------------------------
_CANONICAL = [
    "N1","N2","N3","N4",
    "S1","S2","S3","S4",
    "E1","E2",
    "W1","W2",
    "C1","C2",
    "NE1","NE2",
    "X1","X2","X3",
]

ZONE_EXPANSION: Dict[str, List[str]] = {
    # Simplified regional labels
    "N":            ["N1","N2","N3","N4"],
    "NORTH":        ["N1","N2","N3","N4"],
    "S":            ["S1","S2","S3","S4"],
    "SOUTH":        ["S1","S2","S3","S4"],
    "E":            ["E1","E2"],
    "EAST":         ["E1","E2"],
    "W":            ["W1","W2"],
    "WEST":         ["W1","W2"],
    "C":            ["C1","C2"],
    "CENTRAL":      ["C1","C2"],
    "NE":           ["NE1","NE2"],
    "NORTHEAST":    ["NE1","NE2"],
    "NORTH EAST":   ["NE1","NE2"],
    "NE/JK":        ["NE1","NE2","X3"],
    # Special zones
    "J&K":          ["X3"],
    "JK":           ["X3"],
    "LADAKH":       ["X3"],
    "ANDAMAN":      ["X1"],
    "A&N":          ["X1"],
    "LAKSHADWEEP":  ["X2"],
    # 9-zone simplified B2B schemes
    "ZONE A":       ["N1","N2"],
    "ZONE B":       ["N3","N4"],
    "ZONE C":       ["E1","E2"],
    "ZONE D":       ["W1","W2"],
    "ZONE E":       ["S1","S2","S3","S4"],
    "ZONE F":       ["C1","C2"],
    "ZONE G":       ["NE1","NE2"],
    # Extra aliases
    "ZONE 1":       ["N1","N2"],
    "ZONE 2":       ["N3","N4"],
    "ZONE 3":       ["E1","E2"],
    "ZONE 4":       ["W1","W2"],
    "ZONE 5":       ["S1","S2","S3","S4"],
    "METRO":        ["N1","S1","E1","W1"],
    "REST":         ["N2","N3","N4","S2","S3","S4","E2","W2","C1","C2"],
    # Additional common abbreviations used by Indian transporters
    "NCR":          ["N1"],
    "DELHI":        ["N1"],
    "MUMBAI":       ["W1"],
    "CHENNAI":      ["S1"],
    "KOLKATA":      ["E1"],
    "BANGALORE":    ["S2"],
    "BENGALURU":    ["S2"],
    "HYDERABAD":    ["S3"],
    "PUNE":         ["W1"],
    "AHMEDABAD":    ["W1"],
    # Common letter codes used in simplified zone schemes
    "ZA":           ["N1","N2"],
    "ZB":           ["N3","N4"],
    "ZC":           ["E1","E2"],
    "ZD":           ["W1","W2"],
    "ZE":           ["S1","S2","S3","S4"],
    "ZF":           ["C1","C2"],
    "ZG":           ["NE1","NE2"],
    # ROI (Rest of India)
    "ROI":          ["N2","N3","N4","S2","S3","S4","E2","W2","C1","C2","NE1","NE2"],
    "REST OF INDIA":["N2","N3","N4","S2","S3","S4","E2","W2","C1","C2","NE1","NE2"],
    # A/B/C/D zone schemes
    "A":            ["N1","N2"],
    "B":            ["S1","S2"],
    "D":            ["W1","W2"],
}

# Add every canonical zone mapping to itself
for _z in _CANONICAL:
    ZONE_EXPANSION[_z] = [_z]

ZONE_TOKENS = set(ZONE_EXPANSION.keys())

# ---------------------------------------------------------------------------
# Column name sets for pincode list detection
# ---------------------------------------------------------------------------

# Columns where Y = served normally, N = ODA/restricted
_DELIVERY_POSITIVE_NAMES = {
    "delivery", "deliverable", "delivered", "del",
    "delivery status", "is deliverable", "is serviceable",
    "serviceable", "serviceability", "service",
    "served", "is served", "active",
}

# Columns where Y = ODA/not served, N = served (reversed polarity)
_ODA_FLAG_NAMES = {
    "oda", "is oda", "is_oda", "edl", "out of delivery",
    "out_of_delivery", "non serviceable", "non-serviceable",
    "restricted", "oda area", "oda zone",
}

_PICKUP_NAMES = {
    "pickup", "pick up", "pu", "is pickup", "is_pickup",
    "pickup available",
}

_COD_NAMES = {
    "cod", "cash on delivery", "is cod", "is_cod",
    "cod available", "cod serviceable", "cod status",
}

_STATE_NAMES = {
    "state", "state code", "statecode", "state_code", "st",
    "state name",
}

# ODA positive-value strings (explicit ODA column)
ODA_POSITIVE = {"yes", "y", "true", "1", "oda", "x", "edl", "out", "yes-oda"}
ODA_UNICODE  = {"\u2713", "\u2714", "\u2611", "\u2705"}

# ---------------------------------------------------------------------------
# Charge keyword map — every alias maps to a single FC4 field name
# ---------------------------------------------------------------------------
CHARGE_MAP = {
    # Docket / LR
    "docket":                   "docketCharges",
    "docket charge":            "docketCharges",
    "docket charges":           "docketCharges",
    "docket fee":               "docketCharges",
    "lr":                       "docketCharges",
    "lr charge":                "docketCharges",
    "lr charges":               "docketCharges",
    "lr fee":                   "docketCharges",
    "lorry receipt":            "docketCharges",
    "lorry receipt charge":     "docketCharges",
    # Fuel
    "fuel":                     "fuel",
    "fuel surcharge":           "fuel",
    "fuel%":                    "fuel",
    "fuel %":                   "fuel",
    "fsc":                      "fuel",
    "fuel surcharge %":         "fuel",
    "fs":                       "fuel",
    # Min charges
    "minimum":                  "minCharges",
    "min charge":               "minCharges",
    "min charges":              "minCharges",
    "minimum charge":           "minCharges",
    "minimum charges":          "minCharges",
    "min freight":              "minCharges",
    "minimum freight":          "minCharges",
    "minimum chargeable":       "minCharges",
    "min chargeable":           "minCharges",
    "minimum chargable":        "minCharges",
    "min chargable":            "minCharges",
    "min weight":               "minWeight",
    "minimum weight":           "minWeight",
    "min wt":                   "minWeight",
    # Volumetric
    "divisor":                  "divisor",
    "cft":                      "divisor",
    "cft divisor":              "divisor",
    "volumetric":               "divisor",
    "volumetric divisor":       "divisor",
    "k factor":                 "divisor",
    "kfactor":                  "divisor",
    "cfactor":                  "divisor",
    "vol divisor":              "divisor",
    # Green tax
    "green tax":                "greenTax",
    "green":                    "greenTax",
    "green surcharge":          "greenTax",
    # ROV / Insurance
    "rov":                      "rovCharges",
    "rov%":                     "rovCharges",
    "risk coverage":            "rovCharges",
    "risk of value":            "rovCharges",
    "fov":                      "rovCharges",
    "fov charges":              "rovCharges",
    "owner's risk":             "rovCharges",
    "owners risk":              "rovCharges",
    "insurance":                "insuranceCharges",
    "insurance charges":        "insuranceCharges",
    # ODA
    "oda":                      "odaCharges",
    "oda per kg":               "odaCharges",
    "oda charge":               "odaCharges",
    "oda charges":              "odaCharges",
    "out of delivery":          "odaCharges",
    "out-of-delivery":          "odaCharges",
    "out of delivery area":     "odaCharges",
    "edl charge":               "odaCharges",
    "edl charges":              "odaCharges",
    "special area":             "odaCharges",
    "remote area":              "odaCharges",
    # COD
    "cod":                      "codCharges",
    "cod %":                    "codCharges",
    "cod%":                     "codCharges",
    "cod charges":              "codCharges",
    "cash on delivery":         "codCharges",
    "cod percentage":           "codCharges",
    # Handling
    "handling":                 "handlingCharges",
    "handling charges":         "handlingCharges",
    "handling charge":          "handlingCharges",
    # DACC
    "dacc":                     "daccCharges",
    "dacc charges":             "daccCharges",
    # Misc
    "misc":                     "miscCharges",
    "miscellaneous":            "miscCharges",
    "misc charges":             "miscCharges",
    "other charges":            "miscCharges",
    "idc":                      "miscCharges",
    "indirect cost":            "miscCharges",
    # Topay
    "topay":                    "topayCharges",
    "to pay":                   "topayCharges",
    "topay charges":            "topayCharges",
    "to pay charges":           "topayCharges",
    # DOD
    "dod":                      "dodCharges",
    "dod charges":              "dodCharges",
    "delivery on demand":       "dodCharges",
    # Appointment
    "appointment":              "appointmentCharges",
    "apt":                      "appointmentCharges",
    "appointment charges":      "appointmentCharges",
    "apt charges":              "appointmentCharges",
    "apt_handling":             "appointmentCharges",
    "scheduled delivery":       "appointmentCharges",
    # FM
    "fm":                       "fmCharges",
    "first mile":               "fmCharges",
    "fm charges":               "fmCharges",
    "fm charge":                "fmCharges",
    # Prepaid
    "prepaid":                  "prepaidCharges",
    "prepaid charges":          "prepaidCharges",
    "prepaid charge":           "prepaidCharges",
    # E-way bill
    "eway":                     "ewayCharges",
    "e-way":                    "ewayCharges",
    "eway bill":                "ewayCharges",
    "e-way bill":               "ewayCharges",
    "ewaybill":                 "ewayCharges",
    # Underscore variants (DP World style)
    "apt_handling":             "appointmentCharges",
    "min_lr_charge":            "minCharges",
    "min_chg_wt":               "minWeight",
    "green_tax":                "greenTax",
    "to_pay":                   "topayCharges",
    "fuel_surcharge":           "fuel",
    "fuel_surcahrge":           "fuel",   # common typo
    "fuel surcharge %":         "fuel",
    "fuel surcahrge":           "fuel",   # common typo
    # Extra variations
    "minimum chargable freight":"minCharges",
    "minimum chargable weight": "minWeight",
    "minimum chargeable weight":"minWeight",
    "min chargable weight":     "minWeight",
    "min_weight":               "minWeight",
    "demurrage_charge":         "daccCharges",
    "1cft":                     "divisor",
    "cfactor":                  "divisor",
    "fov - owner's risk":       "rovCharges",
    "fov - owners risk":        "rovCharges",
    # Additional common Indian transporter charge names
    "basic freight":            "minCharges",
    "base freight":             "minCharges",
    "base charge":              "minCharges",
    "base charges":             "minCharges",
    "freight charge":           "minCharges",
    "freight charges":          "minCharges",
    "rate":                     None,          # skip ambiguous
    "rate per kg":              None,          # skip – zone rate, not a surcharge
    "sur charge":               "fuel",
    "surcharge":                "fuel",
    "diesel surcharge":         "fuel",
    "dsc":                      "fuel",
    "fuel adjustment factor":   "fuel",
    "faf":                      "fuel",
    "environmental surcharge":  "greenTax",
    "environmental charge":     "greenTax",
    "risk charge":              "rovCharges",
    "consignment risk":         "rovCharges",
    "value added risk":         "rovCharges",
    "octroi":                   "miscCharges",
    "octroi / entry tax":       "miscCharges",
    "entry tax":                "miscCharges",
    "lr no":                    None,          # skip – not a charge value
    "lr number":                None,
    "single piece":             None,   # skip – not a simple scalar
    "claim settlement":         None,   # skip
    "free storage days":        None,   # skip
    "transit time":             None,   # skip
    "delivery days":            None,   # skip
}

COMPANY_MAP = {
    "company":            "name",
    "company name":       "name",
    "transporter":        "name",
    "transporter name":   "name",
    "name":               "name",
    "gst":                "gstNo",
    "gst no":             "gstNo",
    "gst number":         "gstNo",
    "gstin":              "gstNo",
    "pan":                "panNo",
    "pan no":             "panNo",
    "phone":              "phone",
    "mobile":             "phone",
    "contact":            "phone",
    "email":              "email",
    "address":            "address",
    "city":               "city",
    "state":              "state",
    "pincode":            "contact_pincode",
    "mode":               "transportMode",
    "transport mode":     "transportMode",
    "service type":       "serviceType",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _cell_str(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    if s.endswith(".0") and len(s) > 2:
        try:
            int(s[:-2])
            s = s[:-2]
        except ValueError:
            pass
    return s


def _upper(val) -> str:
    return _cell_str(val).upper()


def _is_zone_token(token: str) -> bool:
    t = token.upper().strip()
    if t in ZONE_TOKENS:
        return True
    # Try normalising hyphens/underscores: "Zone-1" → "ZONE 1", "Zone_A" → "ZONE A"
    t2 = re.sub(r'[-_]', ' ', t)
    if t2 != t and t2 in ZONE_TOKENS:
        return True
    return False


def _count_zone_tokens(row: List) -> int:
    return sum(1 for c in row if _is_zone_token(_cell_str(c)))


def _safe_float(val) -> Optional[float]:
    s = _cell_str(val)
    # Strip currency and unit noise
    s = re.sub(r'(?i)\binr\b\.?\s*', '', s)      # INR / INR.
    s = re.sub(r'(?i)rs\.?\s*', '', s)            # Rs. / Rs
    s = re.sub(r'[₹,]', '', s)                    # ₹ and commas
    s = re.sub(r'%', '', s)                         # percent sign
    s = re.sub(r'(?i)\s*(kgs?|kg|days?|lrs?|pieces?|per\s+\w+|/-)\s*$', '', s)  # trailing units
    s = s.strip()
    if not s or s in ("-", "na", "n/a", "nil", "none", "null", "", "x", "n"):
        return None
    # Take first numeric token if mixed (e.g. "200 or Rs 300", "@4%")
    m = re.search(r'(?:^|[@=:]\s*)(\d+(?:\.\d+)?)', s)
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            pass
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _is_valid_pincode(val) -> bool:
    try:
        pin = int(float(_cell_str(val).replace(",", "")))
        return 100000 <= pin <= 999999
    except (ValueError, TypeError):
        return False


def _parse_pincode(val) -> Optional[int]:
    try:
        pin = int(float(_cell_str(val).replace(",", "")))
        if 100000 <= pin <= 999999:
            return pin
        return None
    except (ValueError, TypeError):
        return None


def _parse_vf_from_row(row: List, start_col: int = 1) -> Dict:
    """
    Parse {v, f} charge pattern from a row starting at start_col.

    Handles patterns like:
      ["ODA", "4%",  "990"]            -> {"v": 4.0,  "f": 990.0}
      ["ODA", "4% or Rs 990"]          -> {"v": 4.0,  "f": 990.0}
      ["ODA", "0.1%", "100 minimum"]   -> {"v": 0.1,  "f": 100.0}
      ["ODA", "750"]                   -> {"f": 750.0}  (only fixed)
      ["Fuel", "20%"]                  -> {"v": 20.0}   (only percent)

    Returns dict with "v" and/or "f" keys, or empty dict if nothing found.
    """
    v_val = None
    f_val = None

    cells = []
    for ci in range(start_col, min(start_col + 4, len(row))):
        c = _cell_str(row[ci])
        if c:
            cells.append(c)

    # Full cell text (joined) for regex on "X% or Rs Y" style
    full_text = " ".join(cells)

    # Pattern: "X% or Rs Y" or "X% / Rs Y"
    m = re.search(r'(\d+(?:\.\d+)?)\s*%\s*(?:or|/|,)?\s*(?:rs|inr|₹)?\s*(\d+(?:\.\d+)?)',
                  full_text, re.IGNORECASE)
    if m:
        return {"v": float(m.group(1)), "f": float(m.group(2))}

    # Scan individual cells
    for cell in cells:
        # Percentage value (could be "4%" or "4" in a % column)
        m_pct = re.search(r'^(\d+(?:\.\d+)?)\s*%$', cell.strip())
        if m_pct and v_val is None:
            v_val = float(m_pct.group(1))
            continue

        # "Rs 990" or "₹990" or "990 minimum"
        m_rs = re.search(r'(?:rs|inr|₹)\s*(\d+(?:\.\d+)?)', cell, re.IGNORECASE)
        if m_rs and f_val is None:
            f_val = float(m_rs.group(1))
            continue

        # Plain number
        n = _safe_float(cell)
        if n is not None:
            # Heuristic: if small (0-100), likely percentage; if large (>100), likely fixed
            if n <= 100 and v_val is None and f_val is None:
                v_val = n
            elif n > 100 and f_val is None:
                f_val = n
            elif v_val is None:
                v_val = n
            elif f_val is None:
                f_val = n

    result = {}
    if v_val is not None:
        result["v"] = v_val
    if f_val is not None:
        result["f"] = f_val
    return result


# ---------------------------------------------------------------------------
# Main parser class
# ---------------------------------------------------------------------------

class ExcelParser(BaseParser):
    SUPPORTED_EXTENSIONS = [".xlsx", ".xls", ".csv", ".tsv"]

    def parse(self, file_path: str) -> Dict[str, Any]:
        ext = os.path.splitext(file_path)[1].lower()
        if ext in (".csv", ".tsv"):
            return self._parse_csv(file_path)
        return self._parse_excel(file_path)

    # ------------------------------------------------------------------
    # Top-level file readers
    # ------------------------------------------------------------------

    def _parse_excel(self, file_path: str) -> Dict[str, Any]:
        fname = os.path.basename(file_path)
        print(f"[ExcelParser] Parsing: {fname}")

        sheets = {}

        try:
            import pandas as pd
            print(f"[ExcelParser] Using pandas to read {fname}")

            xf = pd.ExcelFile(file_path)
            sheet_names = xf.sheet_names
            print(f"[ExcelParser] Sheets found: {sheet_names}")

            for sname in sheet_names:
                df = pd.read_excel(
                    xf, sheet_name=sname,
                    header=None,
                    dtype=str,
                    keep_default_na=False,
                )
                rows = []
                for _, row in df.iterrows():
                    r = [_cell_str(v) for v in row]
                    if any(c for c in r):
                        rows.append(r)
                sheets[sname] = rows
                print(f"[Excel:{sname}] Raw rows loaded: {len(rows)}")

        except Exception as pandas_err:
            print(f"[ExcelParser] pandas failed ({pandas_err}), falling back to openpyxl")
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file_path, data_only=True)
                for sname in wb.sheetnames:
                    ws = wb[sname]
                    rows = []
                    for row in ws.iter_rows(values_only=True):
                        r = [_cell_str(c) for c in row]
                        if any(c for c in r):
                            rows.append(r)
                    sheets[sname] = rows
                    print(f"[Excel:{sname}] openpyxl rows loaded: {len(rows)}")
            except Exception as oxl_err:
                print(f"[ExcelParser] openpyxl also failed ({oxl_err})")
                return {"text": "", "tables": [], "data": {}}

        text = ""
        tables = []
        for sname, rows in sheets.items():
            text += f"\n=== Sheet: {sname} ===\n"
            for row in rows[:100]:
                text += "\t".join(row) + "\n"
            tables.append({"name": sname, "rows": rows})

        data = self._auto_detect(sheets)

        return {"text": text, "tables": tables, "data": data}

    def _parse_csv(self, file_path: str) -> Dict[str, Any]:
        fname = os.path.basename(file_path)
        print(f"[ExcelParser] Parsing CSV: {fname}")

        rows = []
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                with open(file_path, "r", encoding=enc) as f:
                    delimiter = "\t" if file_path.endswith(".tsv") else ","
                    reader = csv.reader(f, delimiter=delimiter)
                    rows = [row for row in reader if any(c.strip() for c in row)]
                print(f"[ExcelParser] CSV rows: {len(rows)} (encoding={enc})")
                break
            except (UnicodeDecodeError, Exception):
                rows = []
                continue

        if not rows:
            print(f"[ExcelParser] Could not read CSV {fname}")
            return {"text": "", "tables": [], "data": {}}

        text = "\n".join("\t".join(row) for row in rows[:200])
        sheets = {fname: rows}
        data = self._auto_detect(sheets)
        return {"text": text, "tables": [{"name": fname, "rows": rows}], "data": data}

    # ------------------------------------------------------------------
    # Auto-detection dispatcher
    # ------------------------------------------------------------------

    def _auto_detect(self, sheets: Dict[str, List[List]]) -> Dict:
        """
        Classify each sheet and extract structured data.
        Important: does NOT bail on first match — charges are extracted
        from every sheet regardless of its primary classification.

        Returns dict with keys: zone_matrix, served_pincodes, oda_pincodes,
        zone_pincodes, charges, company_details, _parseAudit
        """
        global _PARSE_AUDIT
        _PARSE_AUDIT = []            # reset for this parse run
        detected: Dict[str, Any] = {}

        for sheet_name, rows in sheets.items():
            if not rows:
                print(f"[Excel:{sheet_name}] Empty sheet — skipping")
                continue

            print(f"[Excel:{sheet_name}] Classifying ({len(rows)} rows) ...")

            classified_as = []

            # 0. OICR: Station-rate format (TCI pincode+rate cards)
            if not detected.get("zone_matrix"):
                try:
                    from parsers.oicr_engine import get_oicr_engine
                    oicr = get_oicr_engine()
                    oicr_result = oicr.detect_station_rate_from_rows(rows, sheet_name)
                    if oicr_result and oicr_result.get("zone_matrix"):
                        detected["zone_matrix"] = oicr_result["zone_matrix"]
                        classified_as.append(
                            f"OICR-STATION-RATE ({len(oicr_result['zone_matrix'])} zones)"
                        )
                        # Also capture pincodes from station data if not yet detected
                        if oicr_result.get("served_pincodes"):
                            detected.setdefault("served_pincodes", [])
                            detected["served_pincodes"].extend(oicr_result["served_pincodes"])
                    elif oicr_result and oicr_result.get("served_pincodes") and not detected.get("served_pincodes"):
                        # OICR got pincodes but no rates — still useful
                        detected.setdefault("served_pincodes", [])
                        detected["served_pincodes"].extend(oicr_result["served_pincodes"])
                except Exception as _oicr_err:
                    print(f"[Excel:{sheet_name}] OICR check failed: {_oicr_err}")

            # 1. Zone matrix (highest value)
            if not detected.get("zone_matrix"):
                zm = self._try_parse_zone_matrix(rows, sheet_name)
                if zm:
                    detected["zone_matrix"] = zm
                    classified_as.append("ZONE MATRIX")

            # 2. Serviceability / pincode list (try always)
            if "zone_matrix" not in classified_as or len(rows) > 100:
                pinlist = self._try_parse_pincode_list(rows, sheet_name)
                if pinlist:
                    detected.setdefault("served_pincodes", [])
                    detected.setdefault("oda_pincodes", [])
                    detected["served_pincodes"].extend(pinlist["served"])
                    detected["oda_pincodes"].extend(pinlist["oda"])
                    if pinlist.get("zone_pincodes"):
                        detected.setdefault("zone_pincodes", {})
                        for zone, pins in pinlist["zone_pincodes"].items():
                            detected["zone_pincodes"].setdefault(zone, [])
                            detected["zone_pincodes"][zone].extend(pins)
                    classified_as.append(
                        f"SERVICEABILITY ({len(pinlist['served'])} pincodes, "
                        f"{len(pinlist['oda'])} ODA)"
                    )

            # 3. Charges — always try (rate cards can have BOTH matrix AND charges)
            charges = self._try_parse_charges(rows, sheet_name)
            if charges:
                detected.setdefault("charges", {})
                # Merge: {v,f} dicts take priority over scalars already found
                for k, v in charges.items():
                    existing = detected["charges"].get(k)
                    if existing is None:
                        detected["charges"][k] = v
                    elif isinstance(v, dict) and not isinstance(existing, dict):
                        detected["charges"][k] = v  # dict wins over scalar
                classified_as.append(f"CHARGES ({list(charges.keys())})")

            # 4. Company info (only if nothing else was found or sheet has few rows)
            if not classified_as or len(rows) < 30:
                company = self._try_parse_company_info(rows, sheet_name)
                if company:
                    detected.setdefault("company_details", {})
                    detected["company_details"].update(company)
                    classified_as.append(f"COMPANY INFO ({list(company.keys())})")

            if classified_as:
                print(f"[Excel:{sheet_name}] -> " + " | ".join(classified_as))
            else:
                print(f"[Excel:{sheet_name}] -> not classified (no recognisable structure)")

        # Deduplicate pincode lists
        if "served_pincodes" in detected:
            before = len(detected["served_pincodes"])
            detected["served_pincodes"] = list(set(detected["served_pincodes"]))
            print(f"[ExcelParser] served_pincodes: {before} -> {len(detected['served_pincodes'])} (after dedup)")
        if "oda_pincodes" in detected:
            before_oda = len(detected["oda_pincodes"])
            detected["oda_pincodes"] = list(set(detected["oda_pincodes"]))
            print(f"[ExcelParser] oda_pincodes: {before_oda} -> {len(detected['oda_pincodes'])} (after dedup)")

        # ── Geo validation: filter impossible pincodes ─────────────────────────
        gv = _get_gv()
        if gv:
            if detected.get("served_pincodes"):
                clean = []
                bad = []
                for p in detected["served_pincodes"]:
                    if gv.is_valid_format(p):
                        clean.append(p)
                    else:
                        bad.append(p)
                if bad:
                    print(f"[ExcelParser] GeoValidator: removed {len(bad)} invalid-format pincodes "
                          f"(e.g. {bad[:3]})")
                detected["served_pincodes"] = clean

        # ── ZoneResolver: canonicalize vague zone labels ──────────────────────
        if detected.get("zone_pincodes") and gv:
            try:
                from knowledge.smart_matcher import SmartMatcher
                from knowledge.zone_resolver import ZoneResolver
                zr = ZoneResolver(gv, SmartMatcher())

                # Validate and clean impossible pin assignments
                cleaned_zp, zp_issues = zr.validate_and_clean_pincode_assignments(
                    {z: [int(p) for p in pins if gv.is_valid_format(p)]
                     for z, pins in detected["zone_pincodes"].items()},
                    strict=False,
                )
                if zp_issues:
                    print(f"[ExcelParser] Zone-pincode issues ({len(zp_issues)}):")
                    for iss in zp_issues[:10]:
                        print(f"  [!] {iss}")

                # Resolve vague zone labels → canonical zones
                resolved_label = zr.resolve_zone_labels(cleaned_zp, min_coverage=35.0)

                # Re-map zone_pincodes under canonical zone keys
                canonical_zp: Dict[str, List[int]] = {}
                for old_label, pins in cleaned_zp.items():
                    canonical_list = resolved_label.get(old_label, [])
                    if not canonical_list:
                        print(f"[ExcelParser] Could not resolve zone label '{old_label}' "
                              f"— keeping as-is for ZoneMapper to handle")
                        canonical_zp.setdefault(old_label, []).extend(pins)
                    elif len(canonical_list) == 1:
                        canonical_zp.setdefault(canonical_list[0], []).extend(pins)
                    else:
                        # Multiple canonical zones: split by pincodes' actual zones
                        for pin in pins:
                            actual = gv.lookup_zone(pin)
                            if actual and actual in canonical_list:
                                canonical_zp.setdefault(actual, []).append(pin)
                            else:
                                # Fallback: assign to first canonical zone
                                canonical_zp.setdefault(canonical_list[0], []).append(pin)

                detected["zone_pincodes"] = {z: list(set(pins))
                                              for z, pins in canonical_zp.items() if pins}
                print(f"[ExcelParser] Zone labels resolved: "
                      f"{list(detected['zone_pincodes'].keys())}")
            except Exception as e:
                print(f"[ExcelParser] ZoneResolver error: {e}")

        detected["_parseAudit"] = list(_PARSE_AUDIT)
        return detected

    # ------------------------------------------------------------------
    # Zone matrix parser
    # ------------------------------------------------------------------

    def _try_parse_zone_matrix(
        self, rows: List[List], sheet_name: str
    ) -> Optional[Dict]:
        """
        Detect and parse a zone price matrix.
        Scans rows 0-9 for a header with >= 3 zone tokens.
        Supports simplified zone labels via ZONE_EXPANSION.
        """
        if len(rows) < 3:
            return None

        scan_limit = min(10, len(rows))
        header_row_idx = None
        header = []

        for ri in range(scan_limit):
            row = rows[ri]
            score = _count_zone_tokens(row)
            print(f"[Excel:{sheet_name}] Row {ri}: zone-token score={score} "
                  f"(cells: {[_cell_str(c) for c in row[:12]]})")
            if score >= 3:
                header_row_idx = ri
                header = [_upper(c) for c in row]
                print(f"[Excel:{sheet_name}] Zone matrix header at row {ri}: {header[:16]}")
                break

        if header_row_idx is None:
            print(f"[Excel:{sheet_name}] No zone matrix header (no row with >= 3 zone tokens)")
            return None

        # Build column index: col_idx -> list of canonical zones
        # Use SmartMatcher (with ZONE_EXPANSION as fallback) for maximum coverage
        sm = _get_sm()
        col_canonical: Dict[int, List[str]] = {}
        for ci, h in enumerate(header):
            h_stripped = h.strip()
            if not h_stripped:
                continue
            # Try direct ZONE_EXPANSION first (exact, fast); also try hyphen/underscore normalised
            expansion = ZONE_EXPANSION.get(h_stripped) or ZONE_EXPANSION.get(re.sub(r'[-_]', ' ', h_stripped))
            if expansion:
                col_canonical[ci] = expansion
            elif sm:
                # SmartMatcher handles fuzzy/geo matches
                r = sm.match_zone(h_stripped, min_confidence=0.65)
                if r.value:
                    col_canonical[ci] = r.value
                    if r.method != "exact":
                        print(f"[Excel:{sheet_name}]   Smart zone match: '{h_stripped}' "
                              f"-> {r.value} (method={r.method}, conf={r.confidence:.2f})")
                        _record_audit("zone", h_stripped, r.value, r.method, r.confidence, sheet_name)

        if len(col_canonical) < 3:
            print(f"[Excel:{sheet_name}] Not enough zone columns after expansion "
                  f"(need >= 3, got {len(col_canonical)})")
            return None

        print(f"[Excel:{sheet_name}] Zone columns: { {ci: v for ci, v in col_canonical.items()} }")

        matrix: Dict[str, Dict[str, float]] = {}
        data_rows_processed = 0
        data_rows_with_rates = 0

        for row in rows[header_row_idx + 1:]:
            if not row:
                continue

            raw_origin = _upper(row[0]).strip() if row else ""
            if not raw_origin:
                continue

            origin_zones = ZONE_EXPANSION.get(raw_origin) or ZONE_EXPANSION.get(re.sub(r'[-_]', ' ', raw_origin))
            if not origin_zones and sm:
                r = sm.match_zone(raw_origin, min_confidence=0.65)
                if r.value:
                    origin_zones = r.value
                    if r.method != "exact":
                        print(f"[Excel:{sheet_name}]   Smart origin match: '{raw_origin}' "
                              f"-> {r.value} ({r.method})")
                        _record_audit("zone", raw_origin, r.value, r.method, r.confidence, sheet_name)
            if not origin_zones:
                continue

            data_rows_processed += 1
            row_has_rate = False

            for orig_zone in origin_zones:
                if orig_zone not in matrix:
                    matrix[orig_zone] = {}

                for ci, dest_zones in col_canonical.items():
                    if ci >= len(row):
                        continue
                    rate = _safe_float(row[ci])
                    if rate is not None and rate > 0:
                        for dest_zone in dest_zones:
                            matrix[orig_zone][dest_zone] = rate
                        row_has_rate = True

            if row_has_rate:
                data_rows_with_rates += 1

        print(f"[Excel:{sheet_name}] Zone matrix: "
              f"{data_rows_processed} origin rows, "
              f"{data_rows_with_rates} with rates, "
              f"{len(matrix)} origin zones")

        if len(matrix) >= 3:
            for orig in list(matrix.keys())[:4]:
                print(f"[Excel:{sheet_name}]   {orig}: {len(matrix[orig])} dest rates "
                      f"(sample: {dict(list(matrix[orig].items())[:4])})")
            return matrix

        print(f"[Excel:{sheet_name}] Zone matrix discarded: only {len(matrix)} zones (need >= 3)")
        return None

    # ------------------------------------------------------------------
    # Serviceability / pincode list parser
    # ------------------------------------------------------------------

    def _try_parse_pincode_list(
        self, rows: List[List], sheet_name: str
    ) -> Optional[Dict]:
        """
        Parse a pincode serviceability / ODA list.

        Header detection:
          - Looks for a pincode column (pincode/pin/postal/zip)
          - Delivery column: Y=served, N=ODA (from "Delivery", "Serviceable" etc.)
          - ODA flag column: Y=ODA (from "ODA", "EDL", "is_oda" etc.)
          - Zone column, State column for zone assignment
          - Pickup and COD columns as informational

        Logic:
          - All pincodes in the list are treated as served
          - If delivery_col present:
              Delivery=Y -> served
              Delivery=N -> served AND ODA (delivered but with extra charge)
          - If oda_col or oda_flag_col present:
              ODA=Y / is_oda=Y -> also mark as ODA
        """
        scan_limit = min(10, len(rows))
        header_row_idx = None
        pin_col = None
        zone_col = None
        delivery_col = None   # Y=served, N=ODA
        oda_col = None        # explicit ODA (Y=ODA, N=not ODA — reversed polarity)
        oda_flag_col = None   # same as oda_col but detected from column name
        pickup_col = None
        cod_col = None
        state_col = None

        for ri in range(scan_limit):
            row = rows[ri]
            header_lower = [_cell_str(c).lower().strip() for c in row]

            # Find pincode column
            for ci, h in enumerate(header_lower):
                if h in ("pincode", "pin", "pin code", "postal", "zip", "zip code",
                         "pin no", "pinno", "postal code"):
                    pin_col = ci
                    break
                if any(kw in h for kw in ("pincode", "postal code", "zip code", "pin code")):
                    pin_col = ci
                    break

            if pin_col is not None:
                header_row_idx = ri
                print(f"[Excel:{sheet_name}] Pincode header at row {ri}: "
                      f"pin_col={pin_col} ('{header_lower[pin_col]}')")

                # Scan same header row for other columns
                for ci, h in enumerate(header_lower):
                    if ci == pin_col:
                        continue
                    if h in _DELIVERY_POSITIVE_NAMES or any(n in h for n in _DELIVERY_POSITIVE_NAMES):
                        delivery_col = ci
                        print(f"[Excel:{sheet_name}] Delivery col: {ci} ('{h}')")
                    elif h in _ODA_FLAG_NAMES or any(n == h for n in _ODA_FLAG_NAMES):
                        oda_flag_col = ci
                        print(f"[Excel:{sheet_name}] ODA-flag col: {ci} ('{h}') — Y=ODA")
                    elif "oda" in h or "edl" in h or "out of delivery" in h:
                        if oda_col is None and oda_flag_col is None:
                            oda_col = ci
                            print(f"[Excel:{sheet_name}] ODA col: {ci} ('{h}')")
                    elif h in _PICKUP_NAMES:
                        pickup_col = ci
                    elif h in _COD_NAMES:
                        cod_col = ci
                    elif h in _STATE_NAMES:
                        state_col = ci
                    if ("zone" in h) and zone_col is None and ci != pin_col:
                        zone_col = ci
                        print(f"[Excel:{sheet_name}] Zone col: {ci} ('{h}')")
                break

        # Auto-detect by value scan if no named header found
        if header_row_idx is None:
            valid_count = sum(
                1 for row in rows[:50] if row and _is_valid_pincode(row[0])
            )
            if valid_count >= 10:
                pin_col = 0
                header_row_idx = -1
                print(f"[Excel:{sheet_name}] Auto-detected pincode column 0 "
                      f"({valid_count}/50 valid in first 50 rows)")
            else:
                return None

        data_start = header_row_idx + 1
        served: List[int] = []
        oda: List[int] = []
        zone_pincodes: Dict[str, List[int]] = {}
        skipped = 0

        has_delivery_col = delivery_col is not None
        has_any_oda_indicator = (oda_col is not None or
                                  oda_flag_col is not None or
                                  has_delivery_col)

        for row in rows[data_start:]:
            if not row or pin_col >= len(row):
                continue
            pin = _parse_pincode(row[pin_col])
            if pin is None:
                skipped += 1
                continue

            is_oda = False

            # --- Delivery column (Y=served normally, N=ODA/restricted) ---
            if delivery_col is not None and delivery_col < len(row):
                del_val = _cell_str(row[delivery_col]).strip().upper()
                if del_val in ("N", "NO", "FALSE", "0", "X", "NOT SERVICEABLE", "NS"):
                    is_oda = True   # delivery=N → ODA
                # Y → normal, no change

            # --- Explicit ODA positive flag (Y=ODA, N=served) ---
            if oda_col is not None and oda_col < len(row):
                oda_raw = _cell_str(row[oda_col]).strip().lower()
                if oda_raw in ODA_POSITIVE or oda_raw in ODA_UNICODE:
                    is_oda = True

            if oda_flag_col is not None and oda_flag_col < len(row):
                oda_raw = _cell_str(row[oda_flag_col]).strip().upper()
                if oda_raw in ("Y", "YES", "TRUE", "1", "X"):
                    is_oda = True

            # All pincodes in the list are served (ODA pincodes are served too —
            # just with extra charges). Only skip if delivery explicitly marked N
            # AND we have NO other ODA context (means truly non-serviceable).
            if delivery_col is not None:
                # With delivery col: N = still served (ODA), Y = served normal
                served.append(pin)
            else:
                served.append(pin)

            if is_oda:
                oda.append(pin)

            # Zone assignment
            if zone_col is not None and zone_col < len(row):
                z_raw = _cell_str(row[zone_col]).strip().upper()
                if z_raw and z_raw not in ("", "NONE", "NULL", "N/A", "NA", "-", "ZONE"):
                    zone_pincodes.setdefault(z_raw, [])
                    zone_pincodes[z_raw].append(pin)
            elif state_col is not None and state_col < len(row):
                # Use state code as zone hint when no zone column
                st = _cell_str(row[state_col]).strip().upper()
                if st:
                    zone_pincodes.setdefault(st, [])
                    zone_pincodes[st].append(pin)

        print(f"[Excel:{sheet_name}] Pincode parse: "
              f"{len(served)} served, {len(oda)} ODA, {skipped} skipped  "
              f"[delivery_col={delivery_col}, oda_col={oda_col}]")

        if zone_pincodes:
            for z, pins in list(zone_pincodes.items())[:5]:
                print(f"[Excel:{sheet_name}]   Zone/State '{z}': {len(pins)} pincodes")
            if len(zone_pincodes) > 5:
                print(f"[Excel:{sheet_name}]   ... and {len(zone_pincodes)-5} more")

        if len(served) < 10:
            print(f"[Excel:{sheet_name}] Too few pincodes ({len(served)}) — discarding")
            return None

        result: Dict[str, Any] = {"served": served, "oda": oda}
        if zone_pincodes:
            result["zone_pincodes"] = zone_pincodes
        return result

    # ------------------------------------------------------------------
    # Charges parser (key-value + {v, f} detection)
    # ------------------------------------------------------------------

    def _try_parse_charges(
        self, rows: List[List], sheet_name: str
    ) -> Optional[Dict]:
        """
        Parse charges from a sheet.

        Patterns handled:
        1. Simple key-value:  "Fuel Surcharge"  |  "20"
        2. Two-value {v,f}:   "ODA"             |  "4%"  |  "990"
        3. Combined string:   "ODA"             |  "4% or Rs 990"
        4. Reversed columns:  ""  |  "ODA"  |  "4%"  |  "990"
        5. ODA weight-band table:
           "ODA Charges" header, then rows with weight ranges + min charges

        Also detects divisor/minWeight as float scalars.
        """
        charges: Dict = {}

        # First pass: ODA structured table detection
        # Try distance×weight matrix first (most specific), then weight-band
        oda_matrix = self._try_parse_oda_distance_matrix(rows, sheet_name)
        if oda_matrix:
            charges["odaCharges"] = oda_matrix
            print(f"[Excel:{sheet_name}]   ODA distance-weight matrix: "
                  f"{len(oda_matrix.get('matrix', []))} distance bands")
        else:
            oda_bands = self._try_parse_oda_weight_bands(rows, sheet_name)
            if oda_bands:
                charges["odaCharges"] = oda_bands
                print(f"[Excel:{sheet_name}]   ODA weight bands: "
                      f"{len(oda_bands.get('bands', []))} weight bands")

        # Second pass: key-value rows (+ single-cell inline text)
        for ri, row in enumerate(rows):
            # Handle single-cell inline text: "Fuel Surcharge-20%", "Docket charges- Rs. 50/-"
            non_empty = [_cell_str(c) for c in row if _cell_str(c)]
            if len(non_empty) == 1:
                inline = non_empty[0]
                inline_charges = self._parse_inline_charge_text(inline, sheet_name)
                for k, v in inline_charges.items():
                    existing = charges.get(k)
                    if existing is None:
                        charges[k] = v
                    elif isinstance(v, dict) and not isinstance(existing, dict):
                        charges[k] = v  # dict is more informative than scalar
                continue

            if len(row) < 2:
                continue

            # Try col-0 as key, then col-1 as key (if col-0 is empty)
            key_candidates = []
            if _cell_str(row[0]).strip():
                key_candidates.append((0, _cell_str(row[0]).lower().strip()))
            if len(row) > 1 and _cell_str(row[1]).strip() and not _cell_str(row[0]).strip():
                key_candidates.append((1, _cell_str(row[1]).lower().strip()))

            for key_col, key in key_candidates:
                val_start = key_col + 1

                # Layer 1: exact match in CHARGE_MAP (longest pattern first)
                mapped_key = _UNSET = object()
                for pattern in sorted(CHARGE_MAP.keys(), key=len, reverse=True):
                    fc4_key = CHARGE_MAP[pattern]
                    if pattern == key or (len(pattern) > 4 and pattern in key):
                        mapped_key = fc4_key
                        break

                # Layer 2: SmartMatcher (fuzzy + synonym + geo)
                if mapped_key is _UNSET:
                    sm = _get_sm()
                    if sm:
                        r = sm.match_charge(key, min_confidence=0.65)
                        if r.value is not None or r.method != "none":
                            mapped_key = r.value
                            if r.method not in ("exact", "none"):
                                print(f"[Excel:{sheet_name}]   Smart charge match: '{key}' "
                                      f"-> {r.value} (method={r.method}, conf={r.confidence:.2f})")
                                _record_audit("charge", key, r.value, r.method, r.confidence, sheet_name)

                if mapped_key is _UNSET:
                    continue  # not found at all
                if mapped_key is None:
                    continue  # explicitly mapped to None = skip this row

                # Skip if this is the ODA bands key and we already have bands
                if mapped_key == "odaCharges" and isinstance(charges.get("odaCharges"), dict):
                    # Only override if the existing is bands
                    if "bands" in charges.get("odaCharges", {}):
                        continue

                # Collect all non-empty cells to the right of key_col
                # (handles wide Excel layouts where value is in column 6, not column 1)
                val_cells = [_cell_str(c) for c in row[val_start:] if _cell_str(c)]
                if not val_cells:
                    continue

                # Build a virtual 2-element row for vf parsing: [key, val_cells...]
                # (scan from first non-empty cell to the right)
                first_val_col = next(
                    (i for i in range(val_start, len(row)) if _cell_str(row[i])),
                    val_start
                )
                vf = _parse_vf_from_row(row, first_val_col)

                if len(vf) >= 2:
                    # Both v and f found — store as dict
                    result_val = vf
                    print(f"[Excel:{sheet_name}]   charge {mapped_key} = {result_val} "
                          f"(v/f from '{key}' row)")
                elif len(vf) == 1:
                    # Only one value — could be percentage or fixed
                    single = list(vf.values())[0]
                    kind = list(vf.keys())[0]
                    if mapped_key in ("fuel", "divisor", "minWeight") or kind == "v":
                        result_val = single
                    else:
                        result_val = vf  # store as dict even with single key
                    print(f"[Excel:{sheet_name}]   charge {mapped_key} = {result_val} "
                          f"({'pct' if kind == 'v' else 'fixed'} from '{key}')")
                else:
                    # Fall back to plain float from first non-empty value cell
                    plain = next(
                        (_safe_float(c) for c in val_cells if _safe_float(c) is not None),
                        None
                    )
                    if plain is None:
                        continue
                    result_val = plain
                    print(f"[Excel:{sheet_name}]   charge {mapped_key} = {result_val} "
                          f"(scalar from '{key}' / '{val_cells[0]}')")

                # Don't overwrite a dict with a scalar
                existing = charges.get(mapped_key)
                if existing is None:
                    charges[mapped_key] = result_val
                elif isinstance(result_val, dict) and not isinstance(existing, dict):
                    charges[mapped_key] = result_val

        if charges:
            print(f"[Excel:{sheet_name}] Charges found: {list(charges.keys())}")
        return charges if charges else None

    def _parse_inline_charge_text(self, text: str, sheet_name: str) -> Dict:
        """
        Parse a single-cell string that encodes both key and value.

        Examples:
          "Fuel Surcharge-20%"       -> {"fuel": 20.0}
          "IDC-5%"                   -> {"miscCharges": 5.0}
          "Docket charges- Rs. 50/-" -> {"docketCharges": 50.0}
          "Green tax in Delhi: Rs.75"-> {"greenTax": 75.0}
          "Minimum weight- 10 kg"    -> {"minWeight": 10.0}
        """
        result = {}
        text_lower = text.lower().strip()

        # Layer 1: exact substring match in CHARGE_MAP
        fc4_key = None
        for pattern, ck in sorted(CHARGE_MAP.items(), key=lambda x: len(x[0]), reverse=True):
            if ck is None:
                continue
            if pattern in text_lower:
                fc4_key = ck
                break

        # Layer 2: SmartMatcher if no exact match
        if fc4_key is None:
            sm = _get_sm()
            if sm:
                # Try matching just the label portion (before any dash/colon/number)
                label_part = re.split(r'[-:–\d]', text)[0].strip()
                if len(label_part) >= 3:
                    r = sm.match_charge(label_part, min_confidence=0.68)
                    if r.value is not None:
                        fc4_key = r.value
                        if r.method != "exact":
                            print(f"[Excel:{sheet_name}]   Smart inline match: '{label_part}' "
                                  f"-> {r.value} (method={r.method})")
                            _record_audit("charge", label_part, r.value, r.method, r.confidence, sheet_name)

        if fc4_key:
            vf = _parse_vf_from_row([text], start_col=0)
            if not vf:
                nums = re.findall(r'(\d+(?:\.\d+)?)', text)
                if nums:
                    val = float(nums[0])
                    if fc4_key in ("fuel", "divisor") or "%" in text:
                        vf = {"v": val}
                    else:
                        vf = {"f": val}
            if vf:
                if len(vf) == 2:
                    result[fc4_key] = vf
                else:
                    result[fc4_key] = list(vf.values())[0]
                print(f"[Excel:{sheet_name}]   inline charge: {fc4_key} = {result[fc4_key]} "
                      f"from '{text[:60]}'")

        return result

    def _try_parse_oda_distance_matrix(
        self, rows: List[List], sheet_name: str
    ) -> Optional[Dict]:
        """
        Parse ODA charge matrix keyed by distance AND weight — the most common
        real-world structure for road transporters (V Express style):

          ODA LOCATION RATE MATRIX
          Distance (Kms) | 0-100 Kgs | 101-500 Kgs | 501-1000 Kgs | ...
          Up to 25       | NIL       | NIL          | NIL           | ...
          26 - 50        | 800       | 1000         | 1500          | ...
          51 - 100       | 1200      | 1500         | 2200          | ...

        Columns can be at any position (handles wide/merged Excel layouts).
        Returns {"type": "distance_weight_matrix", "matrix": [...]} or None.
        """
        # Step 1: Locate the ODA matrix header row
        matrix_header_idx = None
        for ri, row in enumerate(rows):
            row_text = " ".join(_cell_str(c).upper() for c in row)
            if "ODA" in row_text and ("MATRIX" in row_text or "RATE MATRIX" in row_text):
                matrix_header_idx = ri
                break

        if matrix_header_idx is None:
            return None

        print(f"[Excel:{sheet_name}] ODA matrix header at row {matrix_header_idx}")

        # Step 2: Find weight-column header row (within next 3 rows of header)
        dist_col = None
        weight_cols: List[Tuple[int, int, Optional[int]]] = []  # (col_idx, minKg, maxKg)
        col_header_idx = None

        for ri in range(matrix_header_idx + 1, min(matrix_header_idx + 4, len(rows))):
            row = rows[ri]
            found_weight = False
            temp_dist_col = None
            temp_weight_cols = []

            for ci, cell in enumerate(row):
                cell_str = _cell_str(cell).lower().strip()
                if not cell_str:
                    continue
                if any(kw in cell_str for kw in ("distance", "dist", "km", "kms")):
                    temp_dist_col = ci
                elif "kg" in cell_str:
                    # Parse weight range: "0-100 Kgs", "0- 100 Kgs (Rs.)", "101-500 Kgs."
                    nums = re.findall(r'(\d+)', cell_str)
                    if nums:
                        min_kg = int(nums[0])
                        max_kg = int(nums[1]) if len(nums) >= 2 else None
                        temp_weight_cols.append((ci, min_kg, max_kg))
                        found_weight = True

            if found_weight and len(temp_weight_cols) >= 2:
                dist_col = temp_dist_col
                weight_cols = temp_weight_cols
                col_header_idx = ri
                break

        if not weight_cols or dist_col is None:
            return None

        print(f"[Excel:{sheet_name}] ODA matrix col header at row {col_header_idx}: "
              f"dist_col={dist_col}, weight_cols={weight_cols}")

        # Step 3: Parse distance-band rows
        matrix_entries = []

        for row in rows[col_header_idx + 1:]:
            if dist_col >= len(row):
                continue
            dist_raw = _cell_str(row[dist_col]).strip()
            if not dist_raw:
                continue

            # Parse distance range: "Up to 25", "26 - 50", "51 - 100", "201-300", "> 300"
            dist_nums = re.findall(r'(\d+)', dist_raw)
            if not dist_nums:
                continue

            dist_lower_str = dist_raw.lower()
            if "up to" in dist_lower_str or "upto" in dist_lower_str or len(dist_nums) == 1:
                min_dist = 0
                max_dist = int(dist_nums[0])
            elif "above" in dist_lower_str or ">" in dist_raw or "beyond" in dist_lower_str:
                min_dist = int(dist_nums[0])
                max_dist = None
            else:
                min_dist = int(dist_nums[0])
                max_dist = int(dist_nums[1]) if len(dist_nums) >= 2 else None

            # Read charge for each weight column
            bands = []
            for ci, min_kg, max_kg in weight_cols:
                if ci >= len(row):
                    continue
                charge_raw = _cell_str(row[ci]).strip().upper()
                if charge_raw in ("NIL", "NA", "N/A", "-", "", "FREE"):
                    charge = 0
                else:
                    charge = _safe_float(row[ci]) or 0

                band: Dict = {"minKg": min_kg, "charge": float(charge)}
                if max_kg is not None:
                    band["maxKg"] = max_kg
                bands.append(band)

            if bands:
                entry: Dict = {"minDist": min_dist, "bands": bands}
                if max_dist is not None:
                    entry["maxDist"] = max_dist
                matrix_entries.append(entry)

        if len(matrix_entries) >= 2:
            print(f"[Excel:{sheet_name}] ODA distance matrix: {len(matrix_entries)} distance bands, "
                  f"{len(weight_cols)} weight cols each")
            for e in matrix_entries[:3]:
                charges_sample = [b['charge'] for b in e['bands'][:3]]
                print(f"[Excel:{sheet_name}]   dist {e.get('minDist')}-{e.get('maxDist')}km: "
                      f"charges={charges_sample}")
            return {"type": "distance_weight_matrix", "matrix": matrix_entries}

        return None

    def _try_parse_oda_weight_bands(
        self, rows: List[List], sheet_name: str
    ) -> Optional[Dict]:
        """
        Detect ODA charge weight-band tables of the form:
          ODA Charges
          Lower Limit (kg) | Upper Limit (kg) | Min Rs | Additional/kg
          0                | 500              | 750    | 15
          500              |                  | 900    | 12

        Returns {"type": "weight_band", "bands": [{...}]} or None.
        """
        oda_header_row = None
        for ri, row in enumerate(rows):
            first_cells = " ".join(_cell_str(c).lower() for c in row[:4])
            if "oda" in first_cells and ("charges" in first_cells or "charge" in first_cells):
                oda_header_row = ri
                break

        if oda_header_row is None:
            return None

        # Look for a numeric header row (weight limits) in the next 3 rows
        bands = []
        col_map = {}  # col_idx -> field name

        for ri in range(oda_header_row + 1, min(oda_header_row + 5, len(rows))):
            row = rows[ri]
            row_lower = [_cell_str(c).lower().strip() for c in row]

            # Try to identify columns
            if any("lower" in h or "min kg" in h or "wt" in h or "weight" in h
                   for h in row_lower):
                for ci, h in enumerate(row_lower):
                    if "lower" in h or "from" in h:
                        col_map["minKg"] = ci
                    elif "upper" in h or "to" in h or "max" in h:
                        col_map["maxKg"] = ci
                    elif "min" in h and "rs" in h:
                        col_map["minimum"] = ci
                    elif "additional" in h or "per kg" in h or "rate" in h:
                        col_map["perKg"] = ci
                continue

            # Data row — try to parse
            nums = [_safe_float(c) for c in row[:6]]
            valid_nums = [n for n in nums if n is not None]
            if len(valid_nums) < 2:
                break

            band: Dict = {}
            if col_map:
                for field, ci in col_map.items():
                    if ci < len(row):
                        v = _safe_float(row[ci])
                        if v is not None:
                            band[field] = v
            else:
                # Heuristic: assume [lower_kg, upper_kg, min_charge, per_kg]
                if len(valid_nums) >= 1:
                    band["minKg"] = valid_nums[0]
                if len(valid_nums) >= 2:
                    # If second value looks like a weight limit
                    if valid_nums[1] < 10000:
                        band["maxKg"] = valid_nums[1]
                    else:
                        band["minimum"] = valid_nums[1]
                if len(valid_nums) >= 3:
                    band["minimum"] = valid_nums[2]
                if len(valid_nums) >= 4:
                    band["perKg"] = valid_nums[3]

            if band:
                bands.append(band)

        if bands:
            print(f"[Excel:{sheet_name}] ODA weight bands: {len(bands)} bands found")
            return {"type": "weight_band", "bands": bands}
        return None

    # ------------------------------------------------------------------
    # Company info parser
    # ------------------------------------------------------------------

    def _try_parse_company_info(
        self, rows: List[List], sheet_name: str
    ) -> Optional[Dict]:
        company: Dict[str, str] = {}

        sm = _get_sm()
        for row in rows:
            if len(row) < 2:
                continue
            key = _cell_str(row[0]).lower().strip()
            val = _cell_str(row[1]).strip()
            if not val or val.lower() in ("none", "null", "na", "n/a", "-"):
                continue

            # Layer 1: exact match in COMPANY_MAP
            mapped_key = None
            for pattern, mk in COMPANY_MAP.items():
                if pattern == key:
                    mapped_key = mk
                    break

            # Layer 2: SmartMatcher
            if mapped_key is None and sm:
                r = sm.match_company_field(key, min_confidence=0.7)
                if r.value:
                    mapped_key = r.value
                    if r.method != "exact":
                        print(f"[Excel:{sheet_name}]   Smart company match: '{key}' "
                              f"-> {r.value} (method={r.method})")
                        _record_audit("company", key, r.value, r.method, r.confidence, sheet_name)

            if mapped_key:
                company[mapped_key] = val

        if len(company) >= 2:
            print(f"[Excel:{sheet_name}] Company fields: {list(company.keys())}")
            return company

        return None
